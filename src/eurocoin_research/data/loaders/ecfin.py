"""DG-ECFIN Business and Consumer Surveys loader.

Downloads the DG-ECFIN main indicators file directly from the European
Commission's Economy & Finance portal. This provides survey data with
the EA (changing composition) geo code, which is not available via the
Eurostat SDMX API for these datasets.

The file is a wide-format Excel with columns like:
  EA.INDU, EA.SERV, EA.CONS, EA.RETA, EA.BUIL, EA.ESI, EA.EEI
  EU.INDU, EU.SERV, EU.CONS, ...

Each country/aggregate has 7 indicator columns (INDU, SERV, CONS, RETA,
BUIL, ESI, EEI) with monthly observations as rows.

Source: https://economy-finance.ec.europa.eu/economic-forecast-and-surveys/
        business-and-consumer-surveys/download-business-and-consumer-survey-data/
        time-series_en
"""

from __future__ import annotations

import logging
import tempfile
from datetime import datetime
from pathlib import Path

import polars as pl
import requests

from eurocoin_research.config import SeriesSpec
from eurocoin_research.data.loaders.base import BaseLoader, parse_period

logger = logging.getLogger(__name__)

# The DG-ECFIN main indicators SA file (seasonally adjusted)
# The URL contains a date stamp that changes monthly (e.g., nace2_ecfin_2606)
# We try the latest known URL and fall back to a search if needed.
ECFIN_BASE_URL = (
    "https://ec.europa.eu/economy_finance/db_indicators/surveys/"
    "documents/series/nace2_ecfin_{month_stamp}/main_indicators_sa_nace2.zip"
)

# Column mapping: DG-ECFIN column prefix -> our series ID
# EA = Euro Area (changing composition)
EA_COLUMN_MAP = {
    "EA.INDU": "ICI",          # Manufacturing/Industry Confidence
    "EA.SERV": "ServicesCI",   # Services Confidence
    "EA.CONS": "CCI",          # Consumer Confidence
    "EA.RETA": "RetailCI",     # Retail Trade Confidence
    "EA.BUIL": "ConstCI",      # Construction Confidence
    "EA.ESI":  "ESI",          # Economic Sentiment Indicator
}


class ECFINLoader(BaseLoader):
    """Load DG-ECFIN survey data from the direct download file.

    This loader downloads the DG-ECFIN main indicators Excel file,
    extracts the EA (changing composition) columns, and returns them
    as individual series.
    """

    def __init__(
        self,
        base_url: str = "",
        cache_dir: Path | None = None,
        timeout: int = 120,
        month_stamp: str = "2606",  # Latest known stamp (June 2026)
    ) -> None:
        super().__init__(base_url, cache_dir)
        self.timeout = timeout
        self.month_stamp = month_stamp
        self._bulk_data: pl.DataFrame | None = None
        self._tried_stamps: set[str] = set()

    def _get_bulk_data(self) -> pl.DataFrame | None:
        """Download and parse the DG-ECFIN main indicators file.

        Cached in memory after first download.
        """
        if self._bulk_data is not None:
            return self._bulk_data

        import openpyxl
        import zipfile
        import io

        # Try known month stamps (most recent first)
        stamps_to_try = [self.month_stamp, "2606", "2506", "2505", "2504"]
        for stamp in stamps_to_try:
            if stamp in self._tried_stamps:
                continue
            url = ECFIN_BASE_URL.format(month_stamp=stamp)
            try:
                logger.info("Downloading DG-ECFIN indicators from %s...", url)
                resp = requests.get(url, timeout=self.timeout)
                if resp.status_code == 200:
                    self._tried_stamps.add(stamp)
                    # Extract xlsx from zip
                    z = zipfile.ZipFile(io.BytesIO(resp.content))
                    xlsx_names = [n for n in z.namelist() if n.endswith(".xlsx")]
                    if not xlsx_names:
                        logger.warning("No xlsx file found in DG-ECFIN zip")
                        continue
                    xlsx_data = z.read(xlsx_names[0])

                    # Parse xlsx
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        tmp.write(xlsx_data)
                        tmp_path = tmp.name
                    wb = openpyxl.load_workbook(tmp_path, data_only=True)

                    self._bulk_data = self._parse_ecfin_workbook(wb)
                    if self._bulk_data is not None:
                        logger.info(
                            "DG-ECFIN data loaded: %d observations, %d series",
                            len(self._bulk_data),
                            self._bulk_data["series_id"].n_unique(),
                        )
                        return self._bulk_data
            except Exception:
                logger.debug("Failed to download DG-ECFIN data with stamp %s", stamp)
                self._tried_stamps.add(stamp)

        logger.error("Failed to download DG-ECFIN data from any known URL")
        return None

    @staticmethod
    def _parse_ecfin_workbook(wb) -> pl.DataFrame | None:
        """Parse the DG-ECFIN Excel workbook into long-format DataFrame."""
        # Find the MONTHLY sheet
        if "MONTHLY" not in wb.sheetnames:
            logger.error("MONTHLY sheet not found in DG-ECFIN workbook")
            return None

        ws = wb["MONTHLY"]
        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            return None

        # First row = headers (series names like EA.INDU, EU.ESI, etc.)
        # First column = date
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows_data[0])]

        # Find EA columns
        ea_col_indices: dict[int, str] = {}  # col_index -> EA column name
        for i, h in enumerate(headers):
            if h in EA_COLUMN_MAP:
                ea_col_indices[i] = h

        if not ea_col_indices:
            logger.error("No EA columns found in DG-ECFIN data")
            return None

        # Extract data
        records: list[dict] = []
        for row in rows_data[1:]:
            date_val = row[0] if row else None
            if date_val is None:
                continue

            # Parse date (datetime object from Excel)
            if isinstance(date_val, datetime):
                # DG-ECFIN uses end-of-month dates; normalize to first-of-month
                parsed_date = date_val.replace(day=1).date()
            elif isinstance(date_val, str):
                try:
                    dt = datetime.strptime(date_val, "%Y-%m-%d")
                    parsed_date = dt.replace(day=1).date()
                except ValueError:
                    continue
            else:
                continue

            for col_idx, ea_name in ea_col_indices.items():
                value = row[col_idx] if col_idx < len(row) else None
                if value is not None and isinstance(value, (int, float)):
                    records.append({
                        "date": parsed_date,
                        "series_id": EA_COLUMN_MAP[ea_name],
                        "value": float(value),
                    })

        if not records:
            logger.error("No valid observations parsed from DG-ECFIN data")
            return None

        return pl.DataFrame(records)

    def fetch_series(self, spec: SeriesSpec, start: str | None = None) -> pl.DataFrame:
        """Fetch a single survey series from DG-ECFIN bulk data."""
        # Check cache
        cached = self._load_from_cache(spec.id)
        if cached is not None:
            if start:
                start_date = parse_period(start, spec.frequency)
                cached = cached.filter(pl.col("date") >= start_date)
            return cached

        # Load bulk data
        bulk = self._get_bulk_data()
        if bulk is None or len(bulk) == 0:
            logger.warning("DG-ECFIN bulk data unavailable for %s", spec.id)
            return pl.DataFrame(
                schema={"date": pl.Date, "series_id": pl.Utf8, "value": pl.Float64}
            )

        # Extract series by ID
        df = bulk.filter(pl.col("series_id") == spec.id)

        if len(df) == 0:
            logger.warning(
                "Series %s not found in DG-ECFIN data. Available: %s",
                spec.id,
                bulk["series_id"].unique().to_list(),
            )
            return pl.DataFrame(
                schema={"date": pl.Date, "series_id": pl.Utf8, "value": pl.Float64}
            )

        df = df.select(["date", "series_id", "value"]).sort("date")

        # Filter by start date
        if start:
            start_date = parse_period(start, spec.frequency)
            df = df.filter(pl.col("date") >= start_date)

        # Cache
        self._save_to_cache(spec.id, df)
        return df
