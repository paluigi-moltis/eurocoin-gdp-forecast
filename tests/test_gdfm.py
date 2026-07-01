"""Test the GDFM baseline on the real macro panel."""
import sys
sys.path.insert(0, "src")

import logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")

import numpy as np
import polars as pl

from eurocoin_research.config import load_config
from eurocoin_research.data.panel import PanelAssembler
from eurocoin_research.data.target import compute_mlrg, gdp_levels_to_qoq_growth
from eurocoin_research.features.transforms import transform_panel, SERIES_TRANSFORMS
from eurocoin_research.models.gdfm import GDFM

# Load panel
panel = pl.read_parquet("data/processed/panel_monthly.parquet")
print(f"Panel: {len(panel)} months x {len(panel.columns)-1} series")

# Transform to stationary
transformed = transform_panel(panel)
print(f"Transformed panel shape: {transformed.shape}")

# Extract GDP quarterly data for target
gdp_panel = panel.filter(pl.col("GDP_LEVELS").is_not_null()).select(["date", "GDP_LEVELS"])
gdp_levels = gdp_panel["GDP_LEVELS"].to_numpy()
gdp_growth = gdp_levels_to_qoq_growth(gdp_levels)
mlrg = compute_mlrg(gdp_growth, method="ideal", max_lag=8)

print(f"\nGDP growth: {len(gdp_growth)} quarters")
print(f"MLRG: {len(mlrg)} quarters, range [{np.nanmin(mlrg)*100:.2f}%, {np.nanmax(mlrg)*100:.2f}%]")

# Prepare the monthly panel for GDFM (drop GDP, use only monthly covariates)
monthly_cols = [c for c in transformed.columns if c != "date" and c != "GDP_LEVELS"]
monthly_panel = transformed.select(monthly_cols)

# Drop rows with excessive NaN (>50% missing)
arr = monthly_panel.to_numpy().astype(float)
nan_frac = np.isnan(arr).sum(axis=1) / arr.shape[1]
valid_rows = nan_frac < 0.5
arr_clean = arr[valid_rows]
print(f"\nMonthly panel: {arr.shape} -> {arr_clean.shape} (after dropping high-NaN rows)")

# Impute remaining NaNs with column means
col_means = np.nanmean(arr_clean, axis=0)
for j in range(arr_clean.shape[1]):
    mask = np.isnan(arr_clean[:, j])
    arr_clean[mask, j] = col_means[j]

# Estimate GDFM
print("\n=== GDFM Estimation ===")
gdfm = GDFM(n_factors=3, max_lag=12, freq_band=(0.0, np.pi/6))
result = gdfm.fit_transform(arr_clean)

print(f"\nFactors extracted: {result.factors.shape}")
print(f"Eigenvalues: {np.round(result.eigenvalues, 4)}")
print(f"Variance ratio: {result.variance_ratio:.1%}")

# Determine optimal number of factors
optimal_q = gdfm.determine_n_factors(arr_clean, max_factors=8)
print(f"Optimal factors (Bai-Ng): {optimal_q}")

# Project MLRG
# We need to align the monthly factor scores with quarterly GDP target
# For now, aggregate factors to quarterly by averaging 3 months
factor_dates = transformed.filter(valid_rows).select("date").to_series().to_list()
factor_quarterly = []
for i in range(0, len(result.factors), 3):
    chunk = result.factors[i:i+3]
    if len(chunk) > 0:
        factor_quarterly.append(np.mean(chunk, axis=0))
factor_q = np.array(factor_quarterly)

# Align with MLRG (which has same length as GDP)
min_len = min(len(factor_q), len(mlrg))
mlrg_valid = mlrg[:min_len]
factors_valid = factor_q[:min_len]

# Remove NaN from MLRG
mask = ~np.isnan(mlrg_valid)
mlrg_clean = mlrg_valid[mask]
factors_clean = factors_valid[mask]

print(f"\nAligned: {len(mlrg_clean)} quarters (factors {factors_clean.shape})")

# Project
projected = np.full(min_len, np.nan)
w_const = np.column_stack([np.ones(len(factors_clean)), factors_clean])
coeffs, _, _, _ = np.linalg.lstsq(w_const, mlrg_clean, rcond=None)
for i in range(min_len):
    if not np.isnan(mlrg_valid[i]) and i < len(factors_valid):
        if not np.any(np.isnan(factors_valid[i])):
            projected[i] = coeffs[0] + np.dot(coeffs[1:], factors_valid[i])

# Compute correlation
valid_proj = ~np.isnan(projected)
corr = np.corrcoef(projected[valid_proj], mlrg_valid[valid_proj])[0, 1]
rmse = np.sqrt(np.mean((projected[valid_proj] - mlrg_valid[valid_proj])**2))
print(f"\n=== Projection Results ===")
print(f"Correlation with MLRG: {corr:.3f}")
print(f"RMSE: {rmse*100:.4f}%")
print(f"Intercept: {coeffs[0]*100:.4f}%")
print(f"Coefficients: {np.round(coeffs[1:], 4)}")

# Compare with published Eurocoin
print(f"\n=== Comparison with published Eurocoin (pandemic) ===")
import openpyxl
from datetime import datetime
wb = openpyxl.load_workbook("data/Ecoin_realtime.xlsx", data_only=True)
ws = wb["Ecoin_real_time"]
ecoin = {}
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    dt = row[1]
    val = row[2]
    if dt and val and isinstance(dt, datetime):
        q = (dt.month - 1) // 3 + 1
        key = f"{dt.year}-Q{q}"
        if key not in ecoin:
            ecoin[key] = []
        ecoin[key].append(float(val))

gdp_dates = gdp_panel["date"].to_list()
print(f"{'Quarter':<10} {'MLRG':>10} {'GDFM':>10} {'Eurocoin':>10}")
for i, d in enumerate(gdp_dates):
    if d.year in [2019, 2020, 2021, 2022] and i < min_len:
        q = (d.month - 1) // 3 + 1
        key = f"{d.year}-Q{q}"
        mlrg_str = f"{mlrg_valid[i]*100:+.2f}%" if not np.isnan(mlrg_valid[i]) else "N/A"
        gdfm_str = f"{projected[i]*100:+.2f}%" if not np.isnan(projected[i]) else "N/A"
        ecoin_avg = np.mean(ecoin.get(key, [float('nan')]))
        ecoin_str = f"{ecoin_avg:+.2f}" if not np.isnan(ecoin_avg) else "N/A"
        print(f"  {key:<10} {mlrg_str:>10} {gdfm_str:>10} {ecoin_str:>10}")

print("\nGDFM baseline test complete.")
